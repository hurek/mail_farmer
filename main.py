import logging
import os
import re

from dotenv import load_dotenv
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from playwright.sync_api import Playwright, sync_playwright, Page
from twocaptcha import TwoCaptcha

from profile_generator import MailProfile
from utils import MONTHS

load_dotenv()

api_key = os.getenv("TWO_CAPTCHA_KEY")
if api_key is None:
	raise Exception("2Captcha api-key not found. Put your api-key to .env file")

solver = TwoCaptcha(api_key)

logging.basicConfig(format='[%(levelname)s] [%(asctime)s] %(message)s', level=logging.INFO)


def save_profile(mail_profile: MailProfile) -> None:
	filename = 'emails.xlsx'
	try:
		wb = load_workbook(filename)
		worksheet = wb.worksheets[0]
	except FileNotFoundError:
		headers_row = ["Email", "Password", "Username", "Name", "Last Name", "Gender", "Birthday"]
		wb = Workbook()
		worksheet = wb.active
		worksheet.append(headers_row)

	worksheet.append([
		f"{mail_profile.username}{mail_profile.email_suffix}",
		mail_profile.password,
		mail_profile.username,
		mail_profile.name,
		mail_profile.lastname,
		mail_profile.gender,
		mail_profile.dob.strftime("%d-%m-%Y")
	])
	wb.save(filename)
	logging.info(f"Profile {mail_profile.username}{mail_profile.email_suffix} successfully saved!")


def solve_captcha(page: Page):
	frame = page.wait_for_selector("iframe[data-hcaptcha-widget-id]")
	url = frame.get_attribute("src")

	site_key = re.search(r"sitekey=([\w-]+)", url).group(1)
	code = solver.hcaptcha(site_key, page.url)['code']

	page.evaluate(
		"args => args[0].setAttribute('data-hcaptcha-response', args[1])",
		[frame, code]
	)
	page.evaluate(
		"args => document.querySelector(args[0]).value = args[1]",
		["textarea[name=h-captcha-response]", code]
	)
	page.evaluate("code => hcaptcha.submit(code)", code)
	page.wait_for_timeout(500)


def register(page: Page, mail_profile: MailProfile) -> None:
	logging.info("Registration started!")
	page.goto(
		"https://id.rambler.ru/login-20/mail-registration?rname=head&session=false&back=https%3A%2F%2Fwww.rambler.ru"
		"%2F&param=popup&iframeOrigin=https%3A%2F%2Fwww.rambler.ru "
	)

	page.get_by_placeholder("Почта").click()
	page.get_by_placeholder("Почта").type(mail_profile.username)

	page.locator('input[value="@rambler.ru"]').click()
	page.get_by_text(mail_profile.email_suffix).click()

	page.get_by_placeholder("Придумайте пароль").type(mail_profile.password)
	page.get_by_placeholder("Повтор пароля").type(mail_profile.password)

	page.get_by_placeholder("Выберите вопрос").click()
	page.get_by_text("Кличка домашнего животного").click()
	page.locator("#answer").type("doge")

	logging.info("Solving registration captcha...")
	solve_captcha(page)

	page.click("button[type=submit]")

	page.locator("#firstname").type(mail_profile.name)
	page.locator("#lastname").type(mail_profile.lastname)

	page.get_by_placeholder("Выберите пол").click()
	page.get_by_text(mail_profile.gender).click()

	page.get_by_placeholder("День").click()
	page.get_by_text(str(mail_profile.dob.day), exact=True).click()

	page.get_by_placeholder("Месяц").click()
	page.get_by_text(MONTHS[mail_profile.dob.month], exact=True).click()

	page.get_by_placeholder("Год").click()
	page.get_by_text(str(mail_profile.dob.year), exact=True).click()
	page.click("button[type=submit]")
	page.wait_for_timeout(500)


def activate_imap(page: Page) -> None:
	page.goto("https://mail.rambler.ru/settings/mailapps")

	page.wait_for_selector(".rui-Popup-close").click()

	page.get_by_role("button", name="Вкл").click()
	page.wait_for_url("https://mail.rambler.ru/settings/mailapps/change")

	logging.info("Solving settings captcha...")
	solve_captcha(page)

	page.get_by_role("button", name="Отправить").click()
	page.wait_for_timeout(500)
	logging.info("IMAP activated!")


def registration_with_imap(pw: Playwright, mail_profile: MailProfile) -> None:
	browser = pw.chromium.launch(headless=True)
	context = browser.new_context()
	context.add_init_script(path='hcaptcha.js')
	page = context.new_page()

	register(page, mail_profile)
	activate_imap(page)

	profile.registered = True
	context.close()
	browser.close()


if __name__ == "__main__":
	with sync_playwright() as playwright:
		while True:
			profile = MailProfile.generate_new(locale='ru', use_minute_box=False)
			try:
				registration_with_imap(playwright, profile)
				if profile.registered:
					save_profile(profile)
			except Exception:
				logging.exception("Failed to created the profile")

